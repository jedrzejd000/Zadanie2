
import openpyxl as openpyxl
from openpyxl.styles import Border, Side
import pycountry_convert as pc
from datetime import datetime

def loadDataToOneRow(currentRow,currentColumn,value,sheet):
    sheet.cell(row=currentRow, column=currentColumn, value=value)
    adjustColumn(len(value), currentColumn, sheet)
    makeBordertoCell(currentRow, currentColumn, sheet)

def makeBordertoCell(row, col, sheet):
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    sheet.cell(row=row, column=col).border = thin_border

def adjustColumn(stringLenght,currentColumn, sheet):
    current_column_letter = openpyxl.utils.get_column_letter(currentColumn)
    min_column_width = sheet.column_dimensions[current_column_letter].width
    if(min_column_width<stringLenght):
        adjusted_width = stringLenght+1
        sheet.column_dimensions[current_column_letter].width = adjusted_width

def getContinentByCountry(countryName):
    country_alpha2 = pc.country_name_to_country_alpha2(countryName)
    country_continent_code = pc.country_alpha2_to_continent_code(country_alpha2)
    country_continent_name = pc.convert_continent_code_to_continent_name(country_continent_code)
    return country_continent_name


class ExcelCreator:

    def makeFile(self):
        file = openpyxl.Workbook()
        return file

    def makeSheet(self, file):
        sheet = file.active
        return sheet

    def saveFile(self,workbook,filePath):
        workbook.save(filePath)
        print("plik został zapisany w scieżce: " + filePath)


    def makeHeadersBar(self,startColumn,startRow,titlesArray, sheet):
        for i in range(titlesArray.__len__()):
            title = titlesArray[i]
            titleLength = len(title)
            currentColumn=startColumn+i
            currentRow = startRow
            adjustColumn(titleLength,currentColumn,sheet)
            makeBordertoCell(currentRow, currentColumn,sheet)
            sheet.cell(row=startRow, column=currentColumn,value=title)



    def loadDataToRows(self,startColumn, startRow,rowsCount,data,sheet):
        for i in range(rowsCount):
            currentRow = i + startRow
            currentColumn = startColumn
            date_string = data['results'][i]['dob']['date'].split('.')[0]
            date_object = datetime.fromisoformat(date_string)
            date = date_object.strftime("%Y-%m-%d")
            valuesArray=[
            data['results'][i]['name']['first'],
            data['results'][i]['name']['last'],
            data['results'][i]['gender'],
            date,
            data['results'][i]['phone'],
            data['results'][i]['email'],
            data['results'][i]['location']['country']
            ]

            gender = valuesArray[2]
            correctCountry = "Nie spełnia"
            country = valuesArray[6]
            continent = getContinentByCountry(country)
            age = data['results'][i]['dob']['age']
            correctAge = "Nie spełnia"
            freeTest = "Nie"

            for j in range(valuesArray.__len__()):
                loadDataToOneRow(currentRow,currentColumn,valuesArray[j],sheet)
                currentColumn=currentColumn+1

            if(gender=="female"):
                if(20>=age>=18): correctAge="Spełnia"
                elif 40>=age>=30: correctAge="Spełnia"
                if continent=="Europe": correctCountry="Spełnia"
            else:
                if(33 >= age >= 22):
                    correctAge = "Spełnia"
                elif 55 >= age >= 44:
                    correctAge = "Spełnia"
                if continent == "North America": correctCountry = "Spełnia"

            if correctCountry == "Spełnia" and correctAge == "Spełnia":
                freeTest = "Tak"

            loadDataToOneRow(currentRow,currentColumn,freeTest,sheet)
            currentColumn = currentColumn + 1
            loadDataToOneRow(currentRow, currentColumn, correctAge, sheet)
            currentColumn = currentColumn + 1
            loadDataToOneRow(currentRow, currentColumn, correctCountry, sheet)









